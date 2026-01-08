
import io
from typing import List, Type, Dict, Any, Optional, get_origin, get_args, Union
from datetime import date, datetime
from enum import Enum

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from pydantic.fields import FieldInfo

# Import all schemas from seeder/schemas.py
from app.seeder.schemas import (
    # BAT Schemas
    DriverSchema, AddressSchmea, BankAccountSchema, IndividualSchema,
    CorporationSchema, CorporationOwnerSchema, CorporationPayeesSchema,
    MedallionSchema, MedallionLeaseSchema, DealerSchema, VehicleEntity,
    VehicleSchema, VehicleExpensesSchema, VehicleHackupSchema,
    VehicleInspectionSchema, VehicleRegistrationSchema, LeaseSchema,
    DriverLease,
    # BPM Schemas
    CaseStepConfigSchema, StepConfigSchema, FirstStepConfigSchema,
    CaseStepSchema, CaseTypeSchema, CaseStatusSchema,
    UserSchema, RoleSchema,
    # Shared / Enums
    PaymentMode
)

bat_schemas = {
    "drivers": DriverSchema,
    "address": AddressSchmea,
    "bank_accounts": BankAccountSchema,
    "Individual": IndividualSchema,
    "corporation": CorporationSchema,
    "corporation_owner": CorporationOwnerSchema,
    "corporation_payees": CorporationPayeesSchema,
    "medallion": MedallionSchema,
    "mo_lease": MedallionLeaseSchema,
    "dealers": DealerSchema,
    "vehicle_entity": VehicleEntity,
    "vehicles": VehicleSchema,
    "vehicle_expenses": VehicleExpensesSchema,
    "vehicle_hackups": VehicleHackupSchema,
    "vehicle_inspections": VehicleInspectionSchema,
    "vehicle_registration": VehicleRegistrationSchema,
    "leases": LeaseSchema,
    "lease_driver": DriverLease,
}

bpm_schemas = {
    "users": UserSchema,
    "roles": RoleSchema,
    "CaseStepConfig": CaseStepConfigSchema,
    "CaseStepConfigFiles": StepConfigSchema,
    "CaseFirstStepConfig": FirstStepConfigSchema,
    "CaseStep": CaseStepSchema,
    "CaseTypes": CaseTypeSchema,
    "CaseStatus": CaseStatusSchema
}

SYSTEM_MAPPING = {
    "bat": bat_schemas,
    "bpm": bpm_schemas
}

EXCLUDED_FIELDS = {"id", "created_at", "updated_at", "deleted_at"}

def generate_excel_template_for_system(system: str, include_examples: bool = False) -> io.BytesIO:
    """
    Generates an Excel template for the specified system (bat or bpm).
    """
    if system not in SYSTEM_MAPPING:
        raise ValueError(f"Unknown system: {system}. Valid options: {list(SYSTEM_MAPPING.keys())}")

    schemas = SYSTEM_MAPPING[system]
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)

    # 1. Add Instructions Sheet
    create_instructions_sheet(wb, system)

    # 2. Add Schema Sheets
    for name , schema in schemas.items():
        create_sheet_from_schema(wb, name ,schema, include_examples)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_instructions_sheet(wb: openpyxl.Workbook, system: str):
    ws = wb.create_sheet("README", 0)
    ws.sheet_properties.tabColor = "FF0000" # Red tab for visibility
    
    instructions = [
        ("Instruction", "Details"),
        ("Purpose", f"Data Entry Template for {system.upper()} System"),
        ("Mandatory Fields", "Headers in BOLD are mandatory."),
        ("Optional Fields", "Headers in normal font are optional."),
        ("Enums/Dropdowns", "Select values from the dropdown list where available."),
        ("Dates", "Format dates as YYYY-MM-DD (e.g., 2023-12-31)."),
        ("Booleans", "Select TRUE or FALSE from dropdown."),
        ("Do Not Modify", "Do not change sheet names or header text."),
    ]
    
    # Write Header
    ws.append(["Category", "Information"])
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Write Content
    for row in instructions:
        ws.append(row)
        
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

def create_sheet_from_schema(wb: openpyxl.Workbook, name : str ,schema: Type[BaseModel], include_examples: bool = False):
    sheet_name = name
    ws = wb.create_sheet(sheet_name)
    
    # Freeze Header
    ws.freeze_panes = "A2"
    
    # Get Fields
    fields = schema.model_fields
    
    headers = []
    example_row = []
    
    # We iterate fields to define headers and validations
    col_idx = 1
    for name, field in fields.items():
        if name in EXCLUDED_FIELDS:
            continue
            
        headers.append(name)
        cell = ws.cell(row=1, column=col_idx, value=name)
        
        # Style Header
        is_required = field.is_required()
        if is_required:
             cell.font = Font(bold=True)
        
        # Determine Data Validation
        annotation = field.annotation
        
        # Process Type to find underlying Enum or bool
        target_type = annotation
        # Handle Union[T, None] or Optional[T]
        if get_origin(annotation) in (Union, Optional):
            args = get_args(annotation)
            # Filter out NoneType to get the meaningful type
            valid_args = [arg for arg in args if arg is not type(None)]
            if valid_args:
                target_type = valid_args[0]
        
        # Values for example
        example_value = None
        
        # Enum Validation
        if isinstance(target_type, type) and issubclass(target_type, Enum):
            # Create dropdown
            enum_values = [e.value for e in target_type]
            if enum_values:
                example_value = enum_values[0]
                
            # Excel validation formula string limit is 255 usually
            formula = f'"{",".join(str(v) for v in enum_values)}"'
            if len(formula) < 255:
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                # Apply to the whole column (e.g., rows 2 to 1000)
                dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1000")
        
        # Bool Validation
        elif target_type is bool:
            example_value = True
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1000")
            
        # Date Validation
        elif target_type is date:
            example_value = "2023-01-01"
            ws.column_dimensions[get_column_letter(col_idx)].number_format = 'YYYY-MM-DD'
        elif target_type is datetime:
            example_value = "2023-01-01 12:00:00"
            ws.column_dimensions[get_column_letter(col_idx)].number_format = 'YYYY-MM-DD HH:MM:SS'
        elif target_type is int:
            example_value = 123
        elif target_type is float:
             example_value = 99.99
        else:
            example_value = f"Example {name}"

        if include_examples:
            example_row.append(example_value)

        col_idx += 1
        
    # Add example row if requested
    if include_examples and example_row:
        ws.append(example_row)
        
    # Auto width (approximation)
    for col in range(1, col_idx):
        ws.column_dimensions[get_column_letter(col)].width = 20
